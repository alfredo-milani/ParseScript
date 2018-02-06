import os
import random
import sys
from os.path import isfile

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path

import constants
from constants import SystemConstants, FormsiteConstants
from model import *
from utils import *
from view import DataRetrievalUI


class DataRetrievalController(object):
    """
    Controllore per le view che estendono l'interfaccia in view/DataRetrievalUI
    """

    def __init__(self):
        # View reference
        self.__view_instance = None

    @property
    def view_instance(self):
        return self.__view_instance

    @view_instance.setter
    def view_instance(self, value):
        if not isinstance(value, DataRetrievalUI):
            raise TypeError("Incorrect instance")

        self.__view_instance = value

    @staticmethod
    def count_users(data, user_delim):
        """
        Count occurences' number of @user_delim in @data
        :type data: list
        :type user_delim: str
        :rtype: int
        """
        n = 0
        for el in data:
            if el.find(user_delim) != -1:
                n += 1

        return n

    def perform_operation(self, input_file, output_dir, sheet_title=""):
        """
        Perform parsing operation of file @input_file setting sheet title as @sheet_title and
        store the result in @output_dir directory
        :type input_file: str
        :type output_dir: str
        :type sheet_title: str
        :rtype: None
        """
        file_to_parse = Path(input_file)
        if not file_to_parse.is_file():
            self.__view_instance.print_to_user(
                "File '%s' not found" % file_to_parse,
                DataRetrievalUI.Colors.TEXT_COLOR_ERROR
            )
            return
            # sys.exit(EXIT_ERR_FILE)

        # Conversione in base al formato del file di input
        content = Converter.document_to_text(input_file)

        if content is None:
            self.__view_instance.print_to_user(
                "!!! Unknown format for file: %s. Skipping... !!!" % input_file,
                DataRetrievalUI.Colors.TEXT_COLOR_WARNING
            )
            return
            # sys.exit(EXIT_ERR_FILE)
        elif content == Converter.EXT_XLSX:
            self.__view_instance.print_to_user(
                "Note: *.xlsx file will be parsed with custom procedure\n",
                DataRetrievalUI.Colors.TEXT_COLOR_WARNING
            )

            list_of_users = self.parse_users_list_from_xlsx(input_file)
        else:
            # NOTE: filter returns a list in Python 2 and a generator in Python 3
            # False-ish value include: False, None, 0, '', [], () and all others empty containers
            data_list = filter(None, content.split("\n"))

            raw_data_num_users = self.count_users(data_list, constants.NEW_USER)

            list_of_users = self.parse_users_list(data_list)

            if raw_data_num_users != len(list_of_users):
                self.__view_instance.print_to_user(
                    "WARNING:\tUser raw data: %d\tUser parsed: %d.\tCheck if some user missing\n" % (
                        raw_data_num_users,
                        len(list_of_users)
                    ),
                    DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                )

        wb = Workbook()
        ws = wb.active

        # Create name for new sheet
        if sheet_title is None or len(sheet_title) == 0:
            text_to_split = os.path.basename(input_file).split('.')[0]
            text_to_split = Common.replace_unsupported_char(str(text_to_split), ['-', ' '], '_')
            month_list = text_to_split.split("_")
            try:
                if len(month_list) < 3:
                    raise IndexError('Filename non in formato standard: Formsite giorno mese')
                sheet_title = month_list[len(month_list) - 2] + "-" + month_list[len(month_list) - 1][0:3]
            except IndexError:
                sheet_title = text_to_split
        ws.title = sheet_title

        # Add column headings. NB. these must be strings
        ws.append(constants.HEADER_ROW)
        for user in list_of_users:
            ws.append(user.get_list_from_instance())

        # Adding table to sheet
        tab = Table(displayName="Data", ref="A1:J" + str(len(list_of_users) + 1))
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        default_out = os.path.basename(input_file)
        default_out = os.path.splitext(default_out)[0] + Converter.EXT_XLSX
        if output_dir[len(output_dir) - 1] != '/' and output_dir[len(output_dir) - 1] != '\\':
            output_dir += Converter.split_char()
        file_to_save = output_dir + default_out

        if Path(file_to_save).exists():
            response = self.__view_instance.get_user_input_bool(
                "File: %s already exist.\nDo you want to override it?\n" % file_to_save,
                "Type [Yes] / [No]"
            )

            if not response:
                # TODO potrebbe ancora esistere un file con lo stesso nome -> genera codice da funzione hash
                file_to_save = file_to_save[:-len(Converter.EXT_XLSX)] + "_" + \
                               str(random.randint(0, 100000)) + Converter.EXT_XLSX

        wb.save(file_to_save)
        self.__view_instance.print_to_user("<<< File parsed >>>")

    def manage_operation(self, input_data):
        """
        Launch operation depending on user input
        :type input_data: InputData
        :rtype: None
        """
        input_file = input_data.input_file
        output_dir = input_data.output_dir
        sheet_title = input_data.sheet_title
        verbose = input_data.verbose

        self.__view_instance.print_to_user(
            "<<<--- STARTING OPERATION --->>>",
            DataRetrievalUI.Colors.TEXT_COLOR_START_OPERATION
        )

        if len(input_file) != 0:
            self.__view_instance.print_to_user("Input: " + os.path.abspath(input_file))
        if len(output_dir) != 0:
            if not Path(output_dir).is_dir():
                self.__view_instance.print_to_user(
                    "Output directory '%s' not exist. Using: '%s' instead" % (output_dir, SystemConstants.TMP_PATH),
                    DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                )
                input_data.output_dir = SystemConstants.TMP_PATH
                output_dir = input_data.output_dir
        else:
            input_data.output_dir = SystemConstants.TMP_PATH
            output_dir = input_data.output_dir
        self.__view_instance.print_to_user("Output directory: " + os.path.abspath(output_dir))

        if not Path(input_file).exists():
            self.__view_instance.print_to_user(
                "ERROR: %s not exist!" % input_file,
                DataRetrievalUI.Colors.TEXT_COLOR_ERROR
            )
            sys.exit(SystemConstants.EXIT_ERR_ARG)

        if Path(input_file).is_dir():
            list_of_files = [
                input_file + Converter.split_char() + f
                for f in os.listdir(input_file)
                if isfile(input_file + Converter.split_char() + f)
            ]
        elif Path(input_file).is_file():
            list_of_files = [input_file]
        else:
            self.__view_instance.print_to_user(
                "ERROR: %s seems not to be a regular file or directory. Exiting..." % input_file,
                DataRetrievalUI.Colors.TEXT_COLOR_ERROR
            )
            sys.exit(SystemConstants.EXIT_ERR_ARG)

        self.__view_instance.print_to_user("File that will be converted:")
        for el in list_of_files:
            self.__view_instance.print_to_user("\t> %s" % el)

        for el in list_of_files:
            self.__view_instance.print_to_user("\n>>> Parsing file: %s" % el)
            if verbose:
                response = self.__view_instance.get_user_input_bool(
                    "Do you want to continue?",
                    "Type [Yes] / [No]"
                )

                if response:
                    self.perform_operation(el, output_dir, sheet_title)
                else:
                    self.__view_instance.print_to_user("Skipping...")
                    continue
            else:
                self.perform_operation(el, output_dir, sheet_title)

        self.__view_instance.print_to_user(
            "\n<<<--- OPERATION COMPLETED --->>>\n",
            DataRetrievalUI.Colors.TEXT_COLOR_SUCCESS
        )

    @staticmethod
    def check_match(to_match, list_to_check):
        """
        Verifica, anche parziale, del matching tra l'elemento @to_match e gli elementi di @list_to_check
        :type to_match: str
        :type list_to_check: list
        :rtype: int
        """
        for i in range(len(list_to_check)):
            if list_to_check[i] in to_match:
                return i

        return -1

    # noinspection PyArgumentList
    def parse_users_list_v1(self, content):
        """
        Crea una lista di tipo User da @content
        :param content: list
        :return: list
        """
        users_list = []
        i = 0
        while i < len(content):
            # new user found
            if constants.NEW_USER in content[i]:
                name = ""
                surname = ""
                email = ""
                ntel = ""
                scores = []

                i += 1

                # parsing scores
                s = 0
                score_list = [x[0] for x in constants.SCORES_LIST]
                for _ in range(constants.SCORES_NUM):
                    while True:
                        item_position = self.check_match(content[i], score_list)
                        if item_position != -1 or constants.NEW_USER in content[i]:
                            break
                        i += 1

                    if constants.NEW_USER in content[i]:
                        break

                    i += 1
                    if content[i] != constants.SCORE_VAL_NEGATIVE:
                        if content[i] == constants.SCORE_VAL_POSITIVE:
                            scores.append(constants.SCORES_LIST[item_position][1])
                        else:
                            self.__view_instance.print_to_user(
                                "Error parsing value of line: %s.\tValue: %s."
                                "\tPosizione elemento della lista: %d.\n" % (
                                    content[i - 1], content[i], i),
                                DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                            )
                            continue

                    i += 1
                    s += 1

                # parsing credentials
                c = 0
                for _ in range(constants.CREDENTIALS_NUM):
                    while True:
                        item_position = self.check_match(content[i], constants.CREDENTIALS_LIST)
                        if item_position != -1 or constants.NEW_USER in content[i]:
                            break
                        i += 1

                    if constants.NEW_USER in content[i]:
                        break

                    i += 1
                    if constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_NAME:
                        name = content[i]
                    elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_SURNAME:
                        surname = content[i]
                    elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_EMAIL:
                        email = content[i]
                    elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_NTEL:
                        ntel = content[i]

                    i += 1
                    c += 1

                user = User(name, email, surname, ntel, scores)
                users_list.append(user)
                if s != constants.SCORES_NUM or c != constants.CREDENTIALS_NUM:
                    self.__view_instance.print_to_user(
                        "WARNING: Error parsing User: " + user + "\n",
                        DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                    )

                if constants.NEW_USER in content[i]:
                    continue

            i += 1

        return users_list

    # noinspection PyArgumentList
    def parse_users_list_v2(self, content):
        """
        Crea una lista di tipo User da @content
        :type content: list
        :rtype: list
        """
        score_list = [x[0] for x in constants.SCORES_LIST]
        users_list = []
        i = 0
        while i < len(content):
            # New user found
            if constants.NEW_USER in content[i]:
                name = ""
                surname = ""
                email = ""
                ntel = ""
                scores = []

                s, c = 0, 0
                while i + 1 < len(content) and constants.NEW_USER not in content[i + 1] and \
                        i + 2 < len(content) and constants.NEW_USER not in content[i + 2]:
                    i += 1
                    item_position = self.check_match(content[i], score_list)
                    # Item found in scores
                    if item_position != -1:
                        i += 1
                        s += 1
                        if content[i] != constants.SCORE_VAL_NEGATIVE:
                            if content[i] == constants.SCORE_VAL_POSITIVE:
                                scores.append(constants.SCORES_LIST[item_position][1])
                            else:
                                # Per far stampare anche l'utente interessato
                                s -= 1
                                self.__view_instance.print_to_user(
                                    "Error parsing value of line: %s.\tValue: %s."
                                    "\tPosizione elemento della lista: %d." %
                                    (content[i - 1], content[i], i),
                                    DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                                )

                        continue

                    item_position = self.check_match(content[i], constants.CREDENTIALS_LIST)
                    # Item found in credentilas
                    if item_position != -1:
                        i += 1
                        c += 1
                        if constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_NAME:
                            name = content[i]
                        elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_SURNAME:
                            surname = content[i]
                        elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_EMAIL:
                            email = content[i]
                        elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_NTEL:
                            ntel = content[i]

                if len(name) != 0 or len(email) != 0 or len(surname) != 0 or len(ntel) != 0 or len(scores) != 0:
                    user = User(name, email, surname, ntel, scores)
                    users_list.append(user)
                    if s != constants.SCORES_NUM or c != constants.CREDENTIALS_NUM:
                        self.__view_instance.print_to_user(
                            "WARNING: Error parsing User: " + user + " at position: " +
                            str(i) + " / " + str(i + 1) + "\n",
                            DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                        )

            i += 1

        return users_list

    def parse_users_list(self, content):
        """
        Crea una lista di tipo User da @content
        :type content: list
        :rtype: list
        """
        score_list = [x[0] for x in constants.SCORES_LIST]
        users_list = []
        i = 0
        while i < len(content):
            # New user found
            if constants.NEW_USER in content[i]:
                name = ""
                surname = ""
                email = ""
                ntel = ""
                date = ""
                scores = []

                s, c = 0, 0
                while i + 1 < len(content) and constants.NEW_USER not in content[i + 1]:
                    i += 1
                    item_position = self.check_match(content[i], score_list)
                    # Item found in scores
                    if item_position != -1:
                        if i + 1 >= len(content) or constants.NEW_USER in content[i + 1]:
                            break
                        elif self.check_match(content[i + 1], score_list) != -1 or \
                                self.check_match(content[i + 1], constants.CREDENTIALS_LIST) != -1:
                            self.__view_instance.print_to_user(
                                "Unexpected parsing new value even if current is not parsed for user: " + user +
                                "\tPosizione elemento della lista: %d." % i,
                                DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                            )
                            continue

                        i += 1
                        # Check input values
                        if content[i] != constants.SCORE_VAL_NEGATIVE:
                            if content[i] == constants.SCORE_VAL_POSITIVE:
                                scores.append(constants.SCORES_LIST[item_position][1])
                            else:
                                self.__view_instance.print_to_user(
                                    "Error parsing score value for the line: '%s'.\tValue: '%s'.\t"
                                    "Posizione elemento della lista: %d." %
                                    (content[i - 1], content[i], i),
                                    DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                                )
                                continue

                        s += 1
                        continue

                    item_position = self.check_match(content[i], constants.CREDENTIALS_LIST)
                    # Item found in credentials
                    if item_position != -1:
                        if i + 1 >= len(content) or constants.NEW_USER in content[i + 1]:
                            break
                        elif self.check_match(content[i + 1], score_list) != -1 or \
                                self.check_match(content[i + 1], constants.CREDENTIALS_LIST) != -1:
                            self.__view_instance.print_to_user(
                                "Unexpected parsing new value even if current is not parsed for user: " + user +
                                "\tPosizione elemento della lista: %d." % i,
                                DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                            )
                            continue

                        i += 1
                        c += 1
                        if constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_NAME:
                            name = content[i]
                        elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_SURNAME:
                            surname = content[i]
                        elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_EMAIL:
                            email = content[i]
                        elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_NTEL:
                            ntel = content[i]
                        elif constants.CREDENTIALS_LIST[item_position] == constants.CREDENTIAL_DATE:
                            date = content[i]

                if name or email or surname or ntel or scores or date:
                    user = User(name, email, surname, ntel, scores, date)
                    users_list.append(user)
                    if s != constants.SCORES_NUM or c != constants.CREDENTIALS_NUM:
                        # Check non effettuato durante l'assegnazione (nel parsing delle credenziali) dal momento che
                        # la funzione filter elimina tutti i valori Falseish
                        # (primary key as in
                        #  https://fs27.formsite.com/lnisrl/form4/fill?1=6536383b0aff580572ef85e25764f3b2)
                        if not (name and email and surname and ntel and date):
                            self.__view_instance.print_to_user(
                                "Error parsing credential (name, email, surname, phone number empty or date) "
                                "for User: " +
                                user,
                                DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                            )
                        elif s == constants.SCORES_NUM:
                            continue

                        self.__view_instance.print_to_user(
                            "WARNING: The user may have been converted incorrectly: " + user + "\n",
                            DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                        )
                else:
                    self.__view_instance.print_to_user(
                        "WARNING: User with all empty entry at position: " + str(i) + " / " + str(i + 1) + "\n",
                        DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                    )

            i += 1

        return users_list

    @staticmethod
    def get_column_from_xlsx(worksheet, column_index):
        """
        Ritorna i valori della colonna relativa alla lettera @column in una lista
        :type worksheet: Workbook[]
        :type column_index: int
        :rtype: list
        """
        column_list = []
        for row in worksheet.iter_rows():
            value = row[column_index].value
            if value is not None:
                column_list.append(value.encode(Converter.ENCODE_UTF_8, "ignore"))

        return column_list

    @staticmethod
    def get_score_from_column_xlsx(column, score_list):
        """
        Ritorna la lista di interi contenente lo score dell'utente che occupa
        la posizione @column nelle liste contenute in @score_list
        :param column: posizione utente
        :param score_list: lista di liste; ogni lista contenuta in questo parametro rappresenta i
        valori per lo score corrispondente all'i-esimo utente della lista
        :type column: int
        :type score_list: list
        :rtype: list
        """
        score = []
        for i in range(FormsiteConstants.SCORES_NUM):
            if score_list[i][column] == FormsiteConstants.SCORE_VAL_POSITIVE:
                score.append(FormsiteConstants.SCORES_LIST[i][1])

        return score

    @staticmethod
    def get_column_index(column_list, string):
        """
        Ritorna la posizione dell'elementeo @string nella lista @column_list
        :type column_list: list
        :type string: str
        :rtype: int
        """
        for el in range(len(column_list)):
            if string == column_list[el]:
                return el

        return -1

    def parse_users_list_from_xlsx(self, filename):
        """
        Crea una lista di @User da @filename in formato *.xlsx
        :type filename: str
        :rtype: list
        """
        workbook = load_workbook(filename=filename)
        sheet_ranges = workbook.get_sheet_names()
        # Encoding from unicode string (u'string') to utf-8 string
        for sheet in range(len(sheet_ranges)):
            sheet_ranges[sheet] = sheet_ranges[sheet].encode(Converter.ENCODE_UTF_8, "ignore")

        users = []
        for sheet in sheet_ranges:
            worksheet = workbook[sheet]

            header_row = []
            # Get header row
            for column in worksheet.iter_cols():
                # Get only first row's elements
                value = column[0].value
                if value is not None:
                    header_row.append(value.encode(Converter.ENCODE_UTF_8, "ignore"))

            # Get names column
            names_list = self.get_column_from_xlsx(
                worksheet,
                self.get_column_index(header_row, FormsiteConstants.CREDENTIAL_NAME)
            )

            # Get surnames column
            surnames_list = self.get_column_from_xlsx(
                worksheet,
                self.get_column_index(header_row, FormsiteConstants.CREDENTIAL_SURNAME)
            )

            # Get email column
            email_list = self.get_column_from_xlsx(
                worksheet,
                self.get_column_index(header_row, FormsiteConstants.CREDENTIAL_EMAIL)
            )

            # Get ntel column
            ntel_list = self.get_column_from_xlsx(
                worksheet,
                self.get_column_index(header_row, FormsiteConstants.CREDENTIAL_NTEL)
            )

            # Get status column
            status_list = self.get_column_from_xlsx(
                worksheet,
                self.get_column_index(header_row, FormsiteConstants.STATUS_SURVEY)
            )

            # Get end date column
            date_list = self.get_column_from_xlsx(
                worksheet,
                self.get_column_index(header_row, FormsiteConstants.STATUS_SURVEY_DATE)
            )

            # Get score column
            score_list = [
                self.get_column_from_xlsx(
                    worksheet,
                    self.check_match(FormsiteConstants.SCORE_AGE, header_row)
                ),
                self.get_column_from_xlsx(
                    worksheet,
                    self.check_match(FormsiteConstants.SCORE_ALLERGENS, header_row)
                ),
                self.get_column_from_xlsx(
                    worksheet,
                    self.check_match(FormsiteConstants.SCORE_DISTURBANCES, header_row)
                ),
                self.get_column_from_xlsx(
                    worksheet,
                    self.check_match(FormsiteConstants.SCORE_INFECTION, header_row)
                ),
                self.get_column_from_xlsx(
                    worksheet,
                    self.check_match(FormsiteConstants.SCORE_PREGNANT, header_row)
                ),
                self.get_column_from_xlsx(
                    worksheet,
                    self.check_match(FormsiteConstants.SCORE_IMC, header_row)
                )
            ]

            try:
                for i in range(len(status_list)):
                    # Get only users which completed the survey
                    if status_list[i] == FormsiteConstants.STATUS_SURVEY_COMPLETE:
                        users.append(User(
                            names_list[i],
                            email_list[i],
                            surnames_list[i],
                            ntel_list[i],
                            self.get_score_from_column_xlsx(i, score_list),
                            date_list[i]
                        ))
            except IndexError:
                self.__view_instance.print_to_user(
                    "WARNING: Error parsing users at sheet: %s. Probably the document is badly formatted\n" %
                    str(sheet),
                    message_type=DataRetrievalUI.Colors.TEXT_COLOR_WARNING
                )

        return users
