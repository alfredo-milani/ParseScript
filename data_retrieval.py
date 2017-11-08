#!/usr/bin/python
# coding=utf-8
# ============================================================================
# Titolo:           dataRetrieval.py
# Descrizione:      Script per estrapolare dati da un file *.txt ed inserirli in un file *.xlsx
# Autore:           Alfredo Milani (alfredo.milani.94@gmail.com)
# Data:             ven 20 ott 2017, 19.36.00, CEST
# Licenza:          MIT License
# Versione:         0.4.0
# Note:             --/--
# Versione bash:    4.4.12(1)-release
# ============================================================================

# TODO: IMPLEMENTARLO CON CLASSI

import getopt
import os
import platform
import sys

from openpyxl import Workbook
from openpyxl.worksheet.table import TableStyleInfo
from pathlib import Path

import gui_form
import support

EXIT_SUCCESS = 0
EXIT_ERR_ARG = 1
EXIT_ERR_FILE = 2
EXIT_ERR_BAD_CONTENT = 3
EXIT_ERR_PACKAGE_MISSING = 4

LINK_TOOL_CONVERTER_DOCX_TXT = "https://document.online-convert.com/convert-to-txt"
DEFAULT_TMP_WIN = "R:\\"
DEFAULT_TMP_LINUX = "/dev/shm/"
INIT_SCRIPT_NAME = "./utils/scripts/linux/inst_xlsxwriter.sh"
OS_WIN = "Windows"
OS_LINUX = "Linux"
DEFAULT_EXT = ".xlsx"

NEW_USER = "Scoring Summary"
CREDENTIALS_NUM = 4
SCORES_NUM = 6
SCORES_LIST = [
    "1. Hai meno di 18 anni o più di 68?",
    "2. Sei allergico a frutta secca (noci macadamia, anacardi, noci, mandorle, noci pecan), "
    "soia, avena, sesamo, o sedano)?",
    "3. Ti è stato diagnosticato qualche disturbo cronico o assumi farmaci per un qualsiasi "
    "disturbo o patologia, come il Diabete (tipo 1 o tipo 2), patologie cardiovascolari, "
    "renali, epatiche o ha mai avuto episodi di svenimento?",
    "4. Hai frequentemente febbre, tosse, diarrea o segni di infezione attiva?",
    "5. Sei incinta o stai allattando?",
    "6. Hai un indice di massa corporea inferiore (IMC) a 18 o maggiore di 40? "
    "Se non conosci il tuo IMC, che è basato su peso e altezza, clicca qui per trovarlo - clicca qui."
]
CREDENTIALS_LIST = [
    "Nome: *",
    "Cognome: *",
    "Email: *",
    "Telefono: *"
]
HEADER_ROW = [
    "nome",
    "cognome",
    "telefono",
    "email",
    "scores",
    "note"
]


def usage():
    print "\n# Utilizzo\n"
    print "\t./" + os.path.basename(__file__) + " [Options]\n"
    print "# Options\n"
    print "\t-i | --I= | --ifile= )\t\tSetting input file"
    print ("\t-o | --O= | --ofile= )\t\tSetting output file. If not specified, a file (*%s) "
           "will be created in the default temp directory ('%s' -> '%s' | '%s' -> '%s')" %
           (DEFAULT_EXT, OS_WIN, DEFAULT_TMP_WIN, OS_LINUX, DEFAULT_TMP_LINUX))
    print "\t-t | --T= )\t\t\tSetting sheet title. Default behaviour: based on input filename"
    print "\t-h | -H | --help | --HELP )\tShow this help\n"


def set_up_sys():
    try:
        # import xlsxwriter;
        from openpyxl import Workbook
        from openpyxl.compat import range
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as err:
        print str(err)

        os_type = platform.system()
        if os_type == OS_WIN:
            print "TODO: batch script to autosetup package"
            # Required with old library
            # print "Unzip file 'XlsxWriter-RELEASE_1.0.2.zip'"
            # print "Open prompt and execute 'python setup.py install'"
            # print "OR"
            # print "Try with 'pip install packagename' with admin's privileges"
            print "Perform command: 'pip install openpyxl'"
        elif os_type == OS_LINUX:
            print "Perform command: 'sudo pip install openpyxl'"
            # Required with old library
            # subprocess.call([INIT_SCRIPT_NAME])

        sys.exit(EXIT_ERR_PACKAGE_MISSING)


def parse_arg(argv):
    input_file = ""
    output_file = ""
    sheet_title = ""

    try:
        opts, args = getopt.getopt(
            argv,
            "hHt:i:o:",
            ["gui", "GUI", "help", "HELP", "T=", "I=", "O=", "ifile=", "ofile="]
        )
    except getopt.GetoptError as err:
        print str(err)
        usage()
        sys.exit(EXIT_ERR_ARG)

    # print opts;
    # print args;

    if len(opts) == 0 and len(args) == 0:
        usage()
        sys.exit(EXIT_ERR_ARG)

    for opt, arg in opts:
        if opt in ("-h", "-H", "--help", "--HELP"):
            usage()
            sys.exit(EXIT_SUCCESS)
        elif opt in ("-i", "--I", "--ifile"):
            input_file = arg
        elif opt in ("-o", "--O", "--ofile"):
            output_file = arg
        elif opt in ("-t", "--T"):
            sheet_title = arg
        elif opt in ("--gui", "--GUI"):
            gui_form.init()

    print "\nNOTA: Probabilmente a causa di un problema di codifica/decodicifa, i files che mi hai " \
          "dato direttamente in formato *.txt non vengono eleaborati. Per elaborare i dati io ho " \
          "aperto un file in formato *.docx e ho utilizzato la funzione 'salva con nome' e ho " \
          "salvato il file in formato 'Testo (.txt)' (senza codifica)\n"

    print (
        "NOTA: su %s --> utilizza il link '%s' per convertire il file *docx in *txt. "
        "Una volta convertito esegui il comando: "
        "'python X:\script_dir\script.py --I Y:\\file_to_parse.txt --O Z:\create_new_file.xlsx'\n" %
        (OS_WIN, LINK_TOOL_CONVERTER_DOCX_TXT)
    )

    print "Current directory: ", os.getcwd()

    if len(input_file) != 0:
        print "Input file: ", os.path.abspath(input_file)
    if len(output_file) != 0:
        print "Output file: ", os.path.abspath(output_file)
    continue_str = "Do you want to continue? Type [Yes] / [No]\n"
    py3 = sys.version_info[0] > 2
    if py3:
        response = input(continue_str)
    else:
        response = raw_input(continue_str)

    if response == "Yes":
        perform_operation(input_file, output_file, sheet_title)
    else:
        print "Exiting..."
        sys.exit(EXIT_SUCCESS)


def replace_unsupported_char(string, chars_to_check, selected_char):
    for char in chars_to_check:
        string = string.replace(char, selected_char)

    return string


def count_users(data, user_delim):
    n = 0
    for el in data:
        if el.find(user_delim) != -1:
            n += 1

    return n


def perform_operation(input_file, output_file="", sheet_title=""):
    file_to_parse = Path(input_file)
    if not file_to_parse.is_file():
        print ("File '%s' not found" % file_to_parse)
        sys.exit(EXIT_ERR_FILE)

    # with open(input, "r") as file_to_parse:
    #    content = file_to_parse.readlines();

    with open(input_file, "r") as file_to_parse:
        content = file_to_parse.read()
    # content = [x.strip() for x in content];

    data_list = content.split("\n")

    raw_data_num_users = count_users(data_list, NEW_USER)

    data_to_write = parse_list(data_list)

    wb = Workbook()
    ws = wb.active

    if len(sheet_title) == 0:
        text_to_split = os.path.splitext(input_file)[0]
        text_to_split = replace_unsupported_char(text_to_split, ['-', ' '], '_')
        month_list = text_to_split.split("_")
        sheet_title = month_list[1] + "-" + month_list[2][0:3]
    ws.title = sheet_title

    # add column headings. NB. these must be strings
    ws.append(HEADER_ROW)
    for row in data_to_write:
        ws.append(row)

    # tab = Table(displayName="Table1", ref="A1:E200")

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    # tab.tableStyleInfo = style
    # ws.add_table(tab)

    file_to_save = output_file
    default_out = os.path.basename(input_file)
    default_out = os.path.splitext(default_out)[0] + DEFAULT_EXT
    if len(file_to_save) != 0:
        if DEFAULT_EXT not in file_to_save:
            file_to_save += DEFAULT_EXT
    else:
        os_type = platform.system()
        if os_type == OS_WIN:
            file_to_save = DEFAULT_TMP_WIN
        elif os_type == OS_LINUX:
            file_to_save = DEFAULT_TMP_LINUX
        file_to_save += default_out

    wb.save(file_to_save)


def parse_list(content):
    i = 0
    data = []
    # print content;
    # TODO: implementa ricorsione
    content_len = len(content)
    while i < content_len:
        if content[i] == NEW_USER:
            scores = []
            credentials = []

            # TODO incapsulare SCORE_LIST e CREDENTIALA_LIST nell'entità USer
            for itemScore in SCORES_LIST:
                while content[i] != itemScore:
                    i += 1
                i += 1

                if content[i] != "0" and content[i] != "1":
                    print ("Error in line %s" % (content[i - 1]))
                    sys.exit(EXIT_ERR_BAD_CONTENT)

                if content[i] != "0":
                    scores.append(itemScore[0:1])

            for itemCredential in CREDENTIALS_LIST:
                while content[i] != itemCredential:
                    i += 1
                i += 1

                credentials.append(content[i])

            credentials.append(scores)
            data.append(credentials)

        i += 1

    return data


def test_autoconvert():
    # TODO convertitore docx to txt... errore di codifica
    data_to_write = support.document_to_text("/dev/shm/Formsite 20 ottobre.docx")
    f = data_to_write.split("\n")

    # TODO iterator per eliminare elementi in una lista mentre si itera su di essa
    i = 0
    iterator = iter(f)
    while True:
        try:
            if next(iterator) == '':
                del f[i]
            else:
                i += 1
        except StopIteration:
            break

    for el in iterator:
        print el

    input_file = "/dev/shm/Formsite 20 ottobre.docx"
    sheet_title = ""
    output_file = ""

    wb = Workbook()
    ws = wb.active

    if len(sheet_title) == 0:
        text_to_split = os.path.splitext(input_file)[0]
        text_to_split = replace_unsupported_char(text_to_split, ['-', ' '], '_')
        month_list = text_to_split.split("_")
        sheet_title = month_list[1] + "-" + month_list[2][0:3]
    ws.title = sheet_title

    # add column headings. NB. these must be strings
    ws.append(HEADER_ROW)
    for row in f:
        ws.append(list(row))

    file_to_save = output_file
    default_out = os.path.basename(input_file)
    default_out = os.path.splitext(default_out)[0] + DEFAULT_EXT
    if len(file_to_save) != 0:
        if DEFAULT_EXT not in file_to_save:
            file_to_save += DEFAULT_EXT
    else:
        os_type = platform.system()
        if os_type == OS_WIN:
            file_to_save = DEFAULT_TMP_WIN
        elif os_type == OS_LINUX:
            file_to_save = DEFAULT_TMP_LINUX
        file_to_save += default_out

    wb.save(file_to_save)


if __name__ == "__main__":
    set_up_sys()
    parse_arg(sys.argv[1:])
    # test_autoconvert()
