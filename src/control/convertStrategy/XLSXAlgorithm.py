import threading

from openpyxl import load_workbook

from constants import FormsiteConstants
from control.convertStrategy.ConversionAlgorithm import ConversionAlgorithm
from model import User
from utils import Common


class XLSXAlgorithm(ConversionAlgorithm):
    """
    Classe che definisce l'algoritmo per il parsing di documenti in formato *.xlsx
    """

    __instance = None
    __lock = threading.Lock()

    def __init__(self):
        if XLSXAlgorithm.__instance is not None:
            from parsing_exceptions import SingletonException
            raise SingletonException(XLSXAlgorithm)
        else:
            super(XLSXAlgorithm, self).__init__()
            XLSXAlgorithm.__instance = self

    @classmethod
    def get_instance(cls):
        if cls.__instance is None:
            with cls.__lock:
                if cls.__instance is None:
                    XLSXAlgorithm()
        return cls.__instance

    def do_convert(self, file_to_convert):
        from view import ColorsUI
        workbook = load_workbook(filename=file_to_convert)
        sheet_ranges = workbook.sheetnames
        # Encoding from unicode string (u'string') to utf-8 string
        for sheet in range(len(sheet_ranges)):
            sheet_ranges[sheet] = sheet_ranges[sheet].encode(
                ConversionAlgorithm.ENCODE_UTF_8,
                "ignore"
            )

        users = []
        for sheet in sheet_ranges:
            worksheet = workbook[sheet]

            header_row = []
            # Get header row
            for column in worksheet.iter_cols():
                # Get only first row's elements
                value = column[0].value
                if value is not None:
                    header_row.append(value.encode(
                        ConversionAlgorithm.ENCODE_UTF_8,
                        "ignore"
                    ))

            # Get names column
            names_list = self.__get_column_from_xlsx(
                worksheet,
                self.__get_column_index(header_row, FormsiteConstants.CREDENTIAL_NAME)
            )

            # Get surnames column
            surnames_list = self.__get_column_from_xlsx(
                worksheet,
                self.__get_column_index(header_row, FormsiteConstants.CREDENTIAL_SURNAME)
            )

            # Get email column
            email_list = self.__get_column_from_xlsx(
                worksheet,
                self.__get_column_index(header_row, FormsiteConstants.CREDENTIAL_EMAIL)
            )

            # Get ntel column
            ntel_list = self.__get_column_from_xlsx(
                worksheet,
                self.__get_column_index(header_row, FormsiteConstants.CREDENTIAL_NTEL)
            )

            # Get status column
            status_list = self.__get_column_from_xlsx(
                worksheet,
                self.__get_column_index(header_row, FormsiteConstants.STATUS_SURVEY)
            )

            # Get end date column
            date_list = self.__get_column_from_xlsx(
                worksheet,
                self.__get_column_index(header_row, FormsiteConstants.STATUS_SURVEY_DATE)
            )

            # Get score column
            score_list = [
                self.__get_column_from_xlsx(
                    worksheet,
                    Common.check_match(FormsiteConstants.SCORE_AGE, header_row)
                ),
                self.__get_column_from_xlsx(
                    worksheet,
                    Common.check_match(FormsiteConstants.SCORE_ALLERGENS, header_row)
                ),
                self.__get_column_from_xlsx(
                    worksheet,
                    Common.check_match(FormsiteConstants.SCORE_DISTURBANCES, header_row)
                ),
                self.__get_column_from_xlsx(
                    worksheet,
                    Common.check_match(FormsiteConstants.SCORE_INFECTION, header_row)
                ),
                self.__get_column_from_xlsx(
                    worksheet,
                    Common.check_match(FormsiteConstants.SCORE_PREGNANT, header_row)
                ),
                self.__get_column_from_xlsx(
                    worksheet,
                    Common.check_match(FormsiteConstants.SCORE_IMC, header_row)
                )
            ]

            try:
                for i in range(len(status_list)):
                    # Get only users which completed the survey
                    if status_list[i] == FormsiteConstants.STATUS_SURVEY_COMPLETE:
                        users.append(User(
                            names_list[i],
                            email_list[i],
                            surnames_list[i],
                            ntel_list[i],
                            self.__get_score_from_column_xlsx(i, score_list),
                            date_list[i]
                        ))
            except IndexError:
                self.view_instance.print_to_user(
                    "WARNING: Error parsing users at sheet: %s. Probably the document is badly formatted\n" %
                    str(sheet),
                    message_type=ColorsUI.TEXT_COLOR_WARNING
                )

        return users

    @staticmethod
    def __get_column_index(column_list, string):
        """
        Ritorna la posizione dell'elementeo @string nella lista @column_list
        :type column_list: list
        :type string: str
        :rtype: int
        """
        for el in range(len(column_list)):
            if string == column_list[el]:
                return el

        return -1

    @staticmethod
    def __get_column_from_xlsx(worksheet, column_index):
        """
        Ritorna i valori della colonna relativa alla lettera @column in una lista
        :type worksheet: Workbook[]
        :type column_index: int
        :rtype: list
        """
        column_list = []
        for row in worksheet.iter_rows():
            value = row[column_index].value
            if value is not None:
                try:
                    encoded = str(value).encode(ConversionAlgorithm.ENCODE_UTF_8, "ignore")
                except UnicodeEncodeError:
                    encoded = "Encode error, column index: %s" % column_index

                column_list.append(encoded)

        return column_list

    @staticmethod
    def __get_score_from_column_xlsx(column, score_list):
        """
        Ritorna la lista di interi contenente lo score dell'utente che occupa
        la posizione @column nelle liste contenute in @score_list
        :param column: posizione utente
        :param score_list: lista di liste; ogni lista contenuta in questo parametro rappresenta i
        valori per lo score corrispondente all'i-esimo utente della lista
        :type column: int
        :type score_list: list
        :rtype: list
        """
        score = []
        for i in range(FormsiteConstants.SCORES_NUM):
            if score_list[i][column] == FormsiteConstants.SCORE_VAL_POSITIVE:
                score.append(FormsiteConstants.SCORES_LIST[i][1])

        return score
